import tensorflow as tf
import numpy as np
import time
import os
from utils import logger
from automated_reporting import create_training_report

# Try to import openpyxl for Excel export
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
    print("✓ openpyxl is available - Excel export enabled")
except ImportError:
    EXCEL_AVAILABLE = False
    print("✗ WARNING: openpyxl not available. Excel export will not work.")
    print("  Install with: pip install openpyxl")

class Trainer():
    def __init__(self,algo,
                env,
                sampler,
                sample_processor,
                policy,
                n_itr,
                batch_size=500,
                start_itr=0,
                num_inner_grad_steps=3):
        self.algo = algo
        self.env = env
        self.sampler = sampler
        self.sampler_processor = sample_processor
        self.policy = policy
        self.n_itr = n_itr
        self.start_itr = start_itr
        self.num_inner_grad_steps = num_inner_grad_steps
        self.batch_size = batch_size

    def train(self):
        """
        Implement the repilte algorithm for ppo reinforcement learning
        """
        start_time = time.time()
        avg_ret = []
        avg_pg_loss = []
        avg_vf_loss = []

        avg_latencies = []
        avg_greedy_latencies = []  # Track greedy solution latencies
        avg_greedy_energies = []   # Track greedy solution energies (if enabled)
        avg_energies = []          # Track policy energies (if enabled)
        
        # Create directory for detailed Excel exports
        excel_output_dir = "./meta_evaluate_ppo_log/detailed_iterations"
        try:
            os.makedirs(excel_output_dir, exist_ok=True)
            print(f"[DEBUG] Excel output directory created/verified: {excel_output_dir}")
            # Test if directory is writable
            test_file = os.path.join(excel_output_dir, ".test_write")
            try:
                with open(test_file, 'w') as f:
                    f.write("test")
                os.remove(test_file)
                print(f"[DEBUG] Directory is writable")
            except Exception as e:
                print(f"WARNING: Directory may not be writable: {str(e)}")
        except Exception as e:
            print(f"ERROR: Failed to create Excel output directory: {str(e)}")
            excel_output_dir = None
        
        # Energy configuration (same as in your project)
        ENERGY_CONFIG = {
            'rho': 1.0,
            'f_l': 1.0,
            'zeta': 2.0,
            'ptx': 0.1,
            'prx': 0.05,
            'latency_weight': 0.5,
            'energy_weight': 0.5,
            'normalize_energy': True,
        }
        
        for itr in range(self.start_itr, self.n_itr):
            itr_start_time = time.time()
            logger.log("\n ---------------- Iteration %d ----------------" % itr)
            logger.log("Sampling set of tasks/goals for this meta-batch...")

            paths = self.sampler.obtain_samples(log=True, log_prefix='')

            """ ----------------- Processing Samples ---------------------"""
            logger.log("Processing samples...")
            samples_data = self.sampler_processor.process_samples(paths, log='all', log_prefix='')

            """ ------------------- Inner Policy Update --------------------"""
            policy_losses, value_losses = self.algo.UpdatePPOTarget(samples_data, batch_size=self.batch_size)

            #print("task losses: ", losses)
            print("average policy losses: ", np.mean(policy_losses))
            avg_pg_loss.append(np.mean(policy_losses))

            print("average value losses: ", np.mean(value_losses))
            avg_vf_loss.append(np.mean(value_losses))

            """ ------------------- Compute Greedy Solution --------------------"""
            # Compute greedy solution for comparison
            self.env.set_task(0)  # Ensure we're evaluating the correct task
            greedy_result = self.env.greedy_solution()
            
            if self.env.resource_cluster.use_energy:
                greedy_action, greedy_finish_time, greedy_energy = greedy_result
                # Flatten finish times and energy for averaging
                flat_greedy_finish_times = [item for sublist in greedy_finish_time for item in sublist]
                flat_greedy_energy = [item for sublist in greedy_energy for item in sublist]
                avg_greedy_latency = np.mean(flat_greedy_finish_times)
                avg_greedy_energy = np.mean(flat_greedy_energy)
                avg_greedy_latencies.append(avg_greedy_latency)
                avg_greedy_energies.append(avg_greedy_energy)
            else:
                greedy_action, greedy_finish_time = greedy_result
                # Flatten finish times for averaging
                flat_greedy_finish_times = [item for sublist in greedy_finish_time for item in sublist]
                avg_greedy_latency = np.mean(flat_greedy_finish_times)
                avg_greedy_latencies.append(avg_greedy_latency)
                # Calculate greedy energy using method (for consistency)
                avg_greedy_energy = self._calculate_greedy_energy(greedy_action, ENERGY_CONFIG)
                avg_greedy_energies.append(avg_greedy_energy)

            """ ------------------- Logging Stuff --------------------------"""

            ret = np.sum(samples_data['rewards'], axis=-1)
            avg_reward = np.mean(ret)

            latency = samples_data['finish_time']
            avg_latency = np.mean(latency)
            avg_latencies.append(avg_latency)

            # Calculate policy energy consumption
            if self.env.resource_cluster.use_energy:
                avg_energy = self._calculate_policy_energy(samples_data, ENERGY_CONFIG)
                avg_energies.append(avg_energy)
            else:
                avg_energies.append(None)

            logger.logkv('Itr', itr)
            logger.logkv('Average reward, ', avg_reward)
            logger.logkv('Average latency,', avg_latency)
            logger.logkv('Greedy latency,', avg_greedy_latency)
            
            # Log energy if enabled
            if self.env.resource_cluster.use_energy:
                logger.logkv('Average energy,', avg_energy)
                logger.logkv('Greedy energy,', avg_greedy_energy)
                
                # Print energy report after each epoch
                print(f"\n========== EPOCH {itr} ENERGY REPORT ==========")
                print(f"Policy Average Energy: {avg_energy:.6f} Joules")
                print(f"Greedy Average Energy: {avg_greedy_energy:.6f} Joules")
                print(f"Energy Ratio (Policy/Greedy): {avg_energy/avg_greedy_energy:.4f}" if avg_greedy_energy > 0 else "Energy Ratio: N/A")
                print(f"Policy Average Latency: {avg_latency:.6f}")
                print(f"Greedy Average Latency: {avg_greedy_latency:.6f}")
                print(f"===============================================\n")
            
            logger.dumpkvs()
            avg_ret.append(avg_reward)
            
            # Save detailed Excel file for this iteration
            if excel_output_dir is not None:
                try:
                    print(f"\n[DEBUG] Attempting to save Excel for iteration {itr}...")
                    print(f"[DEBUG] EXCEL_AVAILABLE: {EXCEL_AVAILABLE}")
                    print(f"[DEBUG] samples_data keys: {list(samples_data.keys())}")
                    if 'actions' in samples_data:
                        print(f"[DEBUG] actions shape: {samples_data['actions'].shape}")
                    self._save_iteration_excel(itr, samples_data, excel_output_dir)
                    print(f"[DEBUG] Excel save completed for iteration {itr}\n")
                except Exception as e:
                    print(f"ERROR: Failed to save detailed Excel for iteration {itr}: {str(e)}")
                    import traceback
                    traceback.print_exc()
            else:
                print(f"WARNING: Skipping Excel export for iteration {itr} - output directory not available")

        # Generate comprehensive report
        try:
            print("\n==================== GENERATING AUTOMATED REPORT ====================")
            additional_metrics = {
                'policy_losses': avg_pg_loss,
                'value_losses': avg_vf_loss,
                'greedy_latencies': avg_greedy_latencies,
                'average_energy': avg_energies,
                'greedy_energy': avg_greedy_energies
            }
            
            report_dir = create_training_report(
                avg_ret=avg_ret,
                avg_loss=avg_pg_loss,
                avg_latencies=avg_latencies,
                additional_metrics=additional_metrics
            )
            print(f"Report generated successfully at: {report_dir}")
            print("=====================================================================\n")
        except Exception as e:
            print(f"WARNING: Failed to generate automated report: {str(e)}")
            import traceback
            traceback.print_exc()
            print("Training completed successfully but report generation failed.")

        return avg_ret, avg_pg_loss, avg_vf_loss, avg_latencies
    
    def _calculate_policy_energy(self, samples_data, energy_config):
        """
        Calculate energy consumption for policy actions.
        Uses the same energy model as mrlco-new project:
        - Local execution: T_l * rho * (f_l ^ zeta)
        - Offloading (MEC/V2V): T_ul * ptx + T_dl * prx
        """
        # Get actions and finish times
        actions = samples_data['actions']  # Shape: [batch_size, seq_len]
        finish_times = samples_data['finish_time']  # Shape: [batch_size]
        
        total_energy = 0.0
        env = self.env
        
        # Calculate energy for each trajectory
        for i in range(len(finish_times)):
            action_seq = actions[i]
            finish_time = finish_times[i]
            
            # Get the task graph for this trajectory
            task_graph = env.task_graphs_batchs[env.task_id][i % len(env.task_graphs_batchs[env.task_id])]
            
            # Build plan from actions
            plan = []
            for idx, action in enumerate(action_seq):
                if idx < len(task_graph.prioritize_sequence):
                    task_id = task_graph.prioritize_sequence[idx]
                    plan.append((int(task_id), int(action)))
            from env.mec_offloaing_envs.scheduler import schedule_via_adapter

            result, _, _ = schedule_via_adapter(task_graph, plan, env.scheduler_resources)
            total_energy += result.total_mobile_joules
        
        return total_energy / len(finish_times) if len(finish_times) > 0 else 0.0
    
    def _calculate_greedy_energy(self, greedy_action, energy_config):
        """Energy for greedy plans via the canonical scheduler (not the legacy per-task formula)."""
        if not greedy_action or len(greedy_action) == 0:
            return 0.0

        from env.mec_offloaing_envs.scheduler import schedule_via_adapter

        total_energy = 0.0
        env = self.env
        total_plans = 0
        for batch_idx, task_batch in enumerate(greedy_action):
            if batch_idx >= len(env.task_graphs_batchs):
                continue
            task_graphs = env.task_graphs_batchs[batch_idx]
            for plan_idx, plan in enumerate(task_batch):
                if plan_idx >= len(task_graphs):
                    continue
                result, _, _ = schedule_via_adapter(
                    task_graphs[plan_idx], plan, env.scheduler_resources
                )
                total_energy += result.total_mobile_joules
                total_plans += 1
        return total_energy / total_plans if total_plans > 0 else 0.0
    
    def _save_iteration_excel(self, iteration, samples_data, output_dir):
        """
        Save detailed Excel file for each iteration with node-level information.
        
        Args:
            iteration: Current iteration number
            samples_data: Processed sample data containing actions, observations, etc.
            output_dir: Directory to save Excel files
        """
        if not EXCEL_AVAILABLE:
            print(f"ERROR: Skipping Excel export for iteration {iteration}: openpyxl not available")
            print("Please install openpyxl: pip install openpyxl")
            return
        
        print(f"[DEBUG] _save_iteration_excel called for iteration {iteration}")
        print(f"[DEBUG] output_dir: {output_dir}")
        
        # Check if required keys exist
        required_keys = ['actions', 'observations', 'finish_time']
        missing_keys = [key for key in required_keys if key not in samples_data]
        if missing_keys:
            print(f"ERROR: Missing required keys in samples_data: {missing_keys}")
            return
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Iteration_{iteration}"
        
        # Define headers
        headers = [
            'Graph_ID', 'Node_ID', 'Action', 'Action_Name',
            'Latency', 'Total_Time', 'Energy_Consumption', 'Finish_Time',
            'Processing_Data_Size', 'Transmission_Data_Size',
            'Task_Depth', 'Start_Time', 'Execution_Time',
            'Uplink_Time', 'Downlink_Time', 'V2V_Uplink_Time', 'V2V_Downlink_Time',
            'Num_Predecessors', 'Num_Successors'
        ]
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Extract data from samples
        actions = samples_data['actions']  # Shape: [batch_size, seq_len]
        observations = samples_data['observations']  # Shape: [batch_size, seq_len, obs_dim]
        finish_times = samples_data['finish_time']  # Shape: [batch_size]
        
        print(f"[DEBUG] actions shape: {actions.shape}")
        print(f"[DEBUG] observations shape: {observations.shape}")
        print(f"[DEBUG] finish_times shape: {finish_times.shape}")
        print(f"[DEBUG] Number of graphs: {len(actions)}")
        
        # Get energy if available
        energy_data = samples_data.get('energy', None)  # Shape: [batch_size, seq_len] or None
        if energy_data is not None:
            print(f"[DEBUG] energy_data shape: {energy_data.shape if hasattr(energy_data, 'shape') else type(energy_data)}")
        
        row_idx = 2
        env = self.env
        task_graphs_batch = env.task_graphs_batchs[env.task_id]
        print(f"[DEBUG] task_graphs_batch length: {len(task_graphs_batch)}")
        
        # Process each graph (each row in batch corresponds to one graph)
        graphs_processed = 0
        total_nodes_written = 0
        
        for graph_idx, (action_seq, obs_seq, finish_time) in enumerate(zip(actions, observations, finish_times)):
            if graph_idx >= len(task_graphs_batch):
                print(f"[DEBUG] Stopping at graph_idx {graph_idx} (exceeds task_graphs_batch length {len(task_graphs_batch)})")
                break
                
            task_graph = task_graphs_batch[graph_idx]
            graphs_processed += 1
            
            # Get energy for this graph if available
            graph_energy = None
            if energy_data is not None:
                if isinstance(energy_data, np.ndarray):
                    if graph_idx < len(energy_data):
                        graph_energy = energy_data[graph_idx]
                        # If it's a list/array, convert to list for easier indexing
                        if isinstance(graph_energy, np.ndarray):
                            graph_energy = graph_energy.tolist()
                        elif isinstance(graph_energy, (list, tuple)):
                            graph_energy = list(graph_energy)
                elif isinstance(energy_data, (list, tuple)) and graph_idx < len(energy_data):
                    graph_energy = energy_data[graph_idx]
                    if isinstance(graph_energy, np.ndarray):
                        graph_energy = graph_energy.tolist()
            
            # Build plan from actions
            plan = []
            for idx, action in enumerate(action_seq):
                if idx < len(task_graph.prioritize_sequence):
                    task_id = task_graph.prioritize_sequence[idx]
                    plan.append((task_id, int(action)))
            
            # Get detailed scheduling information
            detailed_info = self._get_detailed_scheduling_info(plan, task_graph)
            
            # Write data for each node
            for node_info in detailed_info:
                node_id = node_info['node_id']
                action = node_info['action']
                action_name = ['Local', 'MEC', 'V2V'][action] if 0 <= action <= 2 else 'Unknown'
                
                # Get task properties
                if node_id < len(task_graph.task_list):
                    task = task_graph.task_list[node_id]
                    processing_data_size = task.processing_data_size
                    transmission_data_size = task.transmission_data_size
                    depth = task.depth
                    num_predecessors = len(task_graph.pre_task_sets[node_id])
                    num_successors = len(task_graph.succ_task_sets[node_id])
                else:
                    processing_data_size = 0
                    transmission_data_size = 0
                    depth = 0
                    num_predecessors = 0
                    num_successors = 0
                
                # Get energy for this node (from detailed scheduling info - already calculated correctly)
                node_energy = float(node_info.get('energy', 0.0))
                
                # Calculate total time (including waiting) = finish_time - start_time
                start_time = node_info.get('start_time', 0.0)
                finish_time_node = node_info.get('finish_time', finish_time)
                total_time = max(0.0, finish_time_node - start_time)
                
                # Write row data
                row_data = [
                    graph_idx,  # Graph_ID
                    node_id,  # Node_ID
                    action,  # Action
                    action_name,  # Action_Name
                    node_info.get('latency', 0.0),  # Latency (actual work time)
                    total_time,  # Total_Time (including waiting)
                    node_energy,  # Energy_Consumption
                    finish_time_node,  # Finish_Time
                    processing_data_size,  # Processing_Data_Size
                    transmission_data_size,  # Transmission_Data_Size
                    depth,  # Task_Depth
                    node_info.get('start_time', 0.0),  # Start_Time
                    node_info.get('execution_time', 0.0),  # Execution_Time
                    node_info.get('uplink_time', 0.0),  # Uplink_Time
                    node_info.get('downlink_time', 0.0),  # Downlink_Time
                    node_info.get('v2v_uplink_time', 0.0),  # V2V_Uplink_Time
                    node_info.get('v2v_downlink_time', 0.0),  # V2V_Downlink_Time
                    num_predecessors,  # Num_Predecessors
                    num_successors  # Num_Successors
                ]
                
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    if isinstance(value, (int, float)):
                        cell.number_format = '0.000000'
                
                row_idx += 1
                total_nodes_written += 1
        
        print(f"[DEBUG] Processed {graphs_processed} graphs, wrote {total_nodes_written} node rows")
        
        if total_nodes_written == 0:
            print("WARNING: No rows were written to Excel! Check if actions/observations are empty.")
            return
        
        # Auto-adjust column widths
        for col_idx in range(1, len(headers) + 1):
            max_length = len(headers[col_idx - 1])
            for row_idx in range(2, ws.max_row + 1):
                cell_value = str(ws.cell(row=row_idx, column=col_idx).value)
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 30)
        
        # Save Excel file
        excel_path = os.path.join(output_dir, f"iteration_{iteration}_detailed.xlsx")
        print(f"[DEBUG] Saving Excel to: {excel_path}")
        print(f"[DEBUG] Total rows to write: {row_idx - 2}")
        
        try:
            wb.save(excel_path)
            print(f"✓ Successfully saved detailed Excel for iteration {iteration}: {excel_path} ({row_idx - 2} rows)")
        except Exception as e:
            print(f"ERROR: Failed to save Excel file: {str(e)}")
            import traceback
            traceback.print_exc()
            raise
    
    def _get_detailed_scheduling_info(self, plan, task_graph):
        """Node-level details from the canonical engine (no legacy calendar)."""
        from env.mec_offloaing_envs.scheduler import schedule_via_adapter

        env = self.env
        result, deltas, energy_list = schedule_via_adapter(
            task_graph, plan, env.scheduler_resources
        )
        detailed_info = []
        for action_idx, (node_id, action) in enumerate(plan):
            rec = result.tasks.get(int(node_id))
            if rec is None:
                continue
            hops = [t for t in result.transfers if t.dst_task_id == int(node_id) or t.src_task_id == int(node_id)]
            ul = sum(t.end - t.start for t in hops if t.hop == "MEC_UL")
            dl = sum(t.end - t.start for t in hops if t.hop == "MEC_DL")
            v2v = sum(t.end - t.start for t in hops if t.hop == "V2V")
            node_info = {
                "node_id": int(node_id),
                "action": action,
                "action_idx": action_idx,
                "latency": rec.finish - rec.start,
                "energy": float(energy_list[action_idx]) if action_idx < len(energy_list) else 0.0,
                "finish_time": rec.finish,
                "start_time": rec.start,
                "execution_time": rec.finish - rec.start,
                "uplink_time": ul,
                "downlink_time": dl,
                "v2v_uplink_time": v2v,
                "v2v_downlink_time": 0.0,
                "incremental_makespan": deltas[action_idx] if action_idx < len(deltas) else 0.0,
            }
            detailed_info.append(node_info)
        return detailed_info

if __name__ == "__main__":
    from env.mec_offloaing_envs.offloading_env import Resources
    from env.mec_offloaing_envs.offloading_env import OffloadingEnvironment
    from policies.meta_seq2seq_policy import  Seq2SeqPolicy
    from samplers.seq2seq_sampler import Seq2SeqSampler
    from samplers.seq2seq_sampler_process import Seq2SeSamplerProcessor
    from baselines.vf_baseline import ValueFunctionBaseline
    from meta_algos.ppo_offloading import PPO
    from utils import utils, logger

    logger.configure(dir="./meta_evaluate_ppo_log/task_offloading", format_strs=['stdout', 'log', 'csv'])

    # Energy configuration - enable energy optimization
    ENERGY_CONFIG = {
        'rho': 1.0,           # Local computation energy coefficient
        'f_l': 1.0,           # Local CPU frequency (normalized)
        'zeta': 2.0,          # CPU frequency exponent
        'ptx': 0.1,           # MEC transmission power (Watts)
        'prx': 0.05,          # MEC reception power (Watts)
        'ptx_v2v': 0.06,      # V2V transmission power (Watts, typically < ptx)
        'prx_v2v': 0.03,      # V2V reception power (Watts, typically < prx)
        'rho_v2v': 0.7,       # V2V computation energy coefficient (70% of local)
        'f_v2v': 1.0,         # V2V CPU frequency (normalized, same as local)
        'latency_weight': 0.5, # Weight for latency in combined reward
        'energy_weight': 0.5,  # Weight for energy in combined reward
        'normalize_energy': True,  # Whether to normalize energy rewards
    }

    resource_cluster = Resources(mec_process_capable=(10.0 * 1024 * 1024),
                                 mobile_process_capable=(1.0 * 1024 * 1024),
                                 bandwidth_up=7.0, bandwidth_dl=7.0,
                                 v2v_process_capable=(1.0 * 1024 * 1024),  # Same as UE
                                 v2v_bandwidth=5.0,  # Lower than MEC
                                 use_energy=True,  # Enable energy optimization
                                 energy_config=ENERGY_CONFIG)

    env = OffloadingEnvironment(resource_cluster=resource_cluster,
                                batch_size=100,
                                graph_number=100,
                                graph_file_paths=[
                                    "./env/mec_offloaing_envs/data/meta_offloading_20/offload_random20_12/random.20."
                                    ],
                                time_major=False)

    print("calculate baseline solution======")

    env.set_task(0)
    # Get greedy solution (with energy if enabled)
    greedy_result = env.greedy_solution()
    if env.resource_cluster.use_energy:
        action, finish_time, greedy_energy = greedy_result
        # Flatten finish times and energy for averaging
        flat_finish_times = [item for sublist in finish_time for item in sublist]
        flat_energy = [item for sublist in greedy_energy for item in sublist]
        print("avg greedy solution latency: ", np.mean(flat_finish_times))
        print("avg greedy solution energy: ", np.mean(flat_energy))
    else:
        action, finish_time = greedy_result
        # Flatten finish times for averaging
        flat_finish_times = [item for sublist in finish_time for item in sublist]
        print("avg greedy solution: ", np.mean(flat_finish_times))
    
    # Get reward batch (with energy if enabled)
    reward_result = env.get_reward_batch_step_by_step(action[env.task_id],
                                          env.task_graphs_batchs[env.task_id],
                                          env.max_running_time_batchs[env.task_id],
                                          env.min_running_time_batchs[env.task_id])
    if env.resource_cluster.use_energy:
        target_batch, task_finish_time_batch, energy_batch = reward_result
    else:
        target_batch, task_finish_time_batch = reward_result
    discounted_reward = []
    for reward_path in target_batch:
        discounted_reward.append(utils.discount_cumsum(reward_path, 1.0)[0])

    print("avg greedy solution: ", np.mean(discounted_reward))
    print("avg greedy solution: ", np.mean(task_finish_time_batch))
    print("avg greedy solution: ", np.mean(finish_time))

    print()
    finish_time = env.get_all_mec_execute_time()
    print("avg all remote solution: ", np.mean(finish_time))
    print()
    finish_time = env.get_all_locally_execute_time()
    print("avg all local solution: ", np.mean(finish_time))
    print()
    finish_time = env.get_all_v2v_execute_time()
    print("avg all V2V solution: ", np.mean(finish_time))

    policy = Seq2SeqPolicy(obs_dim=20,
                           encoder_units=128,
                           decoder_units=128,
                           vocab_size=3,
                           name="core_policy")

    sampler = Seq2SeqSampler(env,
                             policy,
                             rollouts_per_meta_task=1,
                             max_path_length=40000,
                             envs_per_task=None,
                             parallel=False)

    baseline = ValueFunctionBaseline()

    sample_processor = Seq2SeSamplerProcessor(baseline=baseline,
                                              discount=0.99,
                                              gae_lambda=0.95,
                                              normalize_adv=True,
                                              positive_adv=False)
    algo = PPO(policy=policy,
               meta_sampler=sampler,
               meta_sampler_process=sample_processor,
               lr=1e-4,
               num_inner_grad_steps=3,
               clip_value=0.2,
               max_grad_norm=None)

    # define the trainer of ppo to evaluate the performance of the trained meta policy for new tasks.
    trainer = Trainer(algo=algo,
                      env=env,
                      sampler=sampler,
                      sample_processor=sample_processor,
                      policy=policy,
                      n_itr=101,
                      start_itr=0,
                      batch_size=500,
                      num_inner_grad_steps=3)

    with tf.Session() as sess:
        sess.run(tf.compat.v1.global_variables_initializer())
        policy.load_variables(load_path="./meta_model_inner_step1/meta_model_final.ckpt")
        avg_ret, avg_pg_loss, avg_vf_loss, avg_latencies = trainer.train()


