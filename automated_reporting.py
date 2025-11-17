"""
Automated Reporting Module for Meta-Trainer Runs
This module creates comprehensive reports after each meta_trainer.run() call.
"""

import os
import json
import csv
import numpy as np
import matplotlib.pyplot as plt
from datetime import datetime
import traceback

# Try to import openpyxl for Excel export (optional)
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Note: openpyxl not available. Excel export will use CSV format (can be opened in Excel).")


class AutomatedReporter:
    """
    Automated reporting system for meta-training runs.
    Creates timestamped folders with data exports, statistics, and plots.
    """
    
    def __init__(self, base_dir="training_reports"):
        """
        Initialize the reporter.
        
        Args:
            base_dir: Base directory for all reports
        """
        self.base_dir = base_dir
        os.makedirs(base_dir, exist_ok=True)
        
    def create_report(self, metrics_data, iteration_data=None):
        """
        Create a comprehensive report for a training run.
        
        Args:
            metrics_data: Dictionary containing metric arrays (e.g., {'rewards': [...], 'losses': [...], ...})
            iteration_data: Optional detailed per-iteration data
            
        Returns:
            report_dir: Path to the created report directory
        """
        try:
            # Step 1: Create timestamped folder
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            report_dir = os.path.join(self.base_dir, timestamp)
            os.makedirs(report_dir)
            print(f"Created report directory: {report_dir}")
            
            # Step 2: Export data
            self._export_data(metrics_data, iteration_data, report_dir)
            
            # Step 3: Export to Excel if available
            if EXCEL_AVAILABLE:
                self._export_to_excel(metrics_data, iteration_data, report_dir)
            
            # Step 4: Compute and save summary statistics
            self._save_summary_statistics(metrics_data, report_dir)
            
            # Step 5: Generate plots
            self._generate_plots(metrics_data, report_dir)
            
            # Step 6: Validation
            self._validate_report(report_dir)
            
            print(f"Report successfully created in: {report_dir}")
            return report_dir
            
        except Exception as e:
            error_msg = f"Failed to create report: {str(e)}\n{traceback.format_exc()}"
            print(f"ERROR: {error_msg}")
            
            # Try to save error log
            try:
                error_file = os.path.join(report_dir if 'report_dir' in locals() else self.base_dir, 
                                        f"error_{timestamp}.txt")
                with open(error_file, 'w') as f:
                    f.write(error_msg)
            except:
                pass
                
            raise
    
    def _export_data(self, metrics_data, iteration_data, report_dir):
        """Export data as CSV and JSON files."""
        # Prepare data for export
        if iteration_data is None:
            # Create iteration data from metrics
            iteration_data = []
            num_iterations = len(next(iter(metrics_data.values())))
            
            for i in range(num_iterations):
                row = {'iteration': i}
                for metric_name, values in metrics_data.items():
                    if i < len(values):
                        row[metric_name] = values[i]
                iteration_data.append(row)
        
        # Save as CSV
        csv_path = os.path.join(report_dir, "data.csv")
        if iteration_data:
            keys = iteration_data[0].keys()
            with open(csv_path, 'w', newline='') as f:
                writer = csv.DictWriter(f, fieldnames=keys)
                writer.writeheader()
                writer.writerows(iteration_data)
            print(f"Saved data.csv with {len(iteration_data)} rows")
        
        # Save as JSON
        json_path = os.path.join(report_dir, "data.json")
        with open(json_path, 'w') as f:
            json.dump({
                'metrics_data': {k: v.tolist() if isinstance(v, np.ndarray) else v 
                               for k, v in metrics_data.items()},
                'iteration_data': iteration_data
            }, f, indent=2)
        print(f"Saved data.json")
    
    def _save_summary_statistics(self, metrics_data, report_dir):
        """Compute and save summary statistics."""
        summary_path = os.path.join(report_dir, "summary.txt")
        
        with open(summary_path, 'w') as f:
            f.write("TRAINING SUMMARY STATISTICS\n")
            f.write("=" * 50 + "\n\n")
            
            for metric_name, values in metrics_data.items():
                if len(values) > 0:
                    values_array = np.array(values)
                    f.write(f"{metric_name.upper()}:\n")
                    f.write(f"  Mean:     {np.mean(values_array):.6f}\n")
                    f.write(f"  Min:      {np.min(values_array):.6f}\n")
                    f.write(f"  Max:      {np.max(values_array):.6f}\n")
                    f.write(f"  Std Dev:  {np.std(values_array):.6f}\n")
                    f.write(f"  Count:    {len(values_array)}\n")
                    f.write("\n")
        
        print(f"Saved summary.txt")
    
    def _export_to_excel(self, metrics_data, iteration_data, report_dir):
        """Export data to Excel format (.xlsx) if openpyxl is available."""
        if not EXCEL_AVAILABLE:
            return
        
        try:
            # Prepare iteration data if not provided
            if iteration_data is None:
                iteration_data = []
                num_iterations = len(next(iter(metrics_data.values())))
                
                for i in range(num_iterations):
                    row = {'iteration': i}
                    for metric_name, values in metrics_data.items():
                        if i < len(values):
                            row[metric_name] = values[i]
                    iteration_data.append(row)
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Training Metrics"
            
            # Write headers
            if iteration_data:
                headers = list(iteration_data[0].keys())
                header_row = 1
                
                # Style headers
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=12)
                
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=header_row, column=col_idx, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Write data
                for row_idx, row_data in enumerate(iteration_data, start=2):
                    for col_idx, header in enumerate(headers, start=1):
                        value = row_data.get(header, "")
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        # Format numeric cells
                        if isinstance(value, (int, float)):
                            cell.number_format = '0.000000'
                
                # Auto-adjust column widths
                for col_idx, header in enumerate(headers, start=1):
                    max_length = len(str(header))
                    for row_idx in range(2, len(iteration_data) + 2):
                        cell_value = str(ws.cell(row=row_idx, column=col_idx).value)
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)
            
            # Save Excel file
            excel_path = os.path.join(report_dir, "data.xlsx")
            wb.save(excel_path)
            print(f"Saved data.xlsx with {len(iteration_data)} rows")
            
        except Exception as e:
            print(f"Warning: Failed to create Excel file: {str(e)}")
            print("CSV file is available and can be opened in Excel.")
    
    def _generate_plots(self, metrics_data, report_dir):
        """Generate plots for all metrics."""
        # Use a matplotlib style that's available in older versions
        try:
            plt.style.use('seaborn-darkgrid')
        except:
            # Fallback to default if seaborn style not available
            plt.style.use('default')
        
        # Define colors for different metric types
        color_map = {
            'energy': '#FF6B6B',  # Red for energy
            'latency': '#4ECDC4',  # Teal for latency
            'reward': '#95E1D3',  # Light green for rewards
            'loss': '#F38181',    # Pink for losses
        }
        
        for metric_name, values in metrics_data.items():
            if len(values) > 0:
                plt.figure(figsize=(10, 6))
                iterations = range(len(values))
                
                # Choose color based on metric name
                color = '#3498db'  # Default blue
                if 'energy' in metric_name.lower():
                    color = color_map.get('energy', '#FF6B6B')
                elif 'latency' in metric_name.lower():
                    color = color_map.get('latency', '#4ECDC4')
                elif 'reward' in metric_name.lower():
                    color = color_map.get('reward', '#95E1D3')
                elif 'loss' in metric_name.lower():
                    color = color_map.get('loss', '#F38181')
                
                plt.plot(iterations, values, linewidth=2, marker='o', markersize=4, 
                        markevery=max(1, len(values)//20), color=color, label=metric_name.replace("_", " ").title())
                
                # Special formatting for energy plots
                if 'energy' in metric_name.lower():
                    plt.title(f'{metric_name.replace("_", " ").title()} Over Iterations (Energy Optimization)', 
                             fontsize=16, fontweight='bold')
                    plt.ylabel('Energy Consumption', fontsize=14)
                else:
                    plt.title(f'{metric_name.replace("_", " ").title()} Over Iterations', 
                             fontsize=16, fontweight='bold')
                    plt.ylabel(metric_name.replace("_", " ").title(), fontsize=14)
                
                plt.xlabel('Iteration', fontsize=14)
                plt.grid(True, alpha=0.3)
                
                # Add trend line
                if len(values) > 1:
                    z = np.polyfit(iterations, values, 1)
                    p = np.poly1d(z)
                    trend_color = 'darkred' if 'energy' in metric_name.lower() else 'r'
                    plt.plot(iterations, p(iterations), "--", alpha=0.8, color=trend_color,
                            label=f'Trend: {z[0]:.2e}x + {z[1]:.2f}')
                
                plt.legend()
                
                # Save plot
                plot_path = os.path.join(report_dir, f"{metric_name}.png")
                plt.tight_layout()
                plt.savefig(plot_path, dpi=150, bbox_inches='tight')
                plt.close()
                
                print(f"Saved {metric_name}.png")
    
    def _validate_report(self, report_dir):
        """Validate that all expected files were created."""
        expected_files = ['data.csv', 'data.json', 'summary.txt']
        
        for filename in expected_files:
            filepath = os.path.join(report_dir, filename)
            if not os.path.exists(filepath):
                raise FileNotFoundError(f"Expected file not found: {filename}")
            
            # Check file is not empty
            if os.path.getsize(filepath) == 0:
                raise ValueError(f"File is empty: {filename}")
        
        # Check for Excel file if available
        if EXCEL_AVAILABLE:
            excel_path = os.path.join(report_dir, "data.xlsx")
            if os.path.exists(excel_path):
                print(f"Excel file (data.xlsx) created successfully")
        
        # Check for at least one plot
        png_files = [f for f in os.listdir(report_dir) if f.endswith('.png')]
        if not png_files:
            raise FileNotFoundError("No plot files (.png) were created")
        
        print(f"Validation passed: All required files present")


def create_training_report(avg_ret, avg_loss, avg_latencies, additional_metrics=None):
    """
    Convenience function to create a report from training metrics.
    
    Args:
        avg_ret: Average returns per iteration
        avg_loss: Average losses per iteration  
        avg_latencies: Average latencies per iteration
        additional_metrics: Optional dictionary of additional metrics
        
    Returns:
        report_dir: Path to created report directory
    """
    reporter = AutomatedReporter()
    
    metrics_data = {
        'average_reward': avg_ret,
        'average_loss': avg_loss,
        'average_latency': avg_latencies
    }
    
    if additional_metrics:
        metrics_data.update(additional_metrics)
    
    return reporter.create_report(metrics_data)


